import json
import math
import os
from collections import Counter, defaultdict
from pathlib import Path
from statistics import median

import openpyxl


ROOT = Path(__file__).resolve().parent
DEFAULT_WORKBOOK = ROOT.parent / "jump_rope" / "venue_2_three_algorithm_accuracy.xlsx"
LEGACY_WORKBOOK = ROOT.parent / "jump_rope" / "\u0032\u53f7\u573a\u5730_\u4e09\u7248\u7b97\u6cd5\u51c6\u786e\u7387.xlsx"
WORKBOOK = Path(
    os.environ.get(
        "DASHBOARD_WORKBOOK",
        DEFAULT_WORKBOOK if DEFAULT_WORKBOOK.exists() or not LEGACY_WORKBOOK.exists() else LEGACY_WORKBOOK,
    )
)
OUT = ROOT / "data.js"
THRESHOLD = 10
VENUE_2 = "Venue 2"

SHEET_ALIASES = (
    "Position Comparison",
    "\u9010\u70b9\u4f4d\u5bf9\u6bd4",
)

COLUMN_ALIASES = {
    "excel_row": ("Excel Row", "\u0045\u0078\u0063\u0065\u006c\u884c"),
    "venue": ("Venue", "\u573a\u5730"),
    "file_name": ("File Name", "\u6587\u4ef6\u540d"),
    "video_id": ("Video ID", "\u89c6\u9891\u0049\u0044"),
    "position": ("Position", "\u70b9\u4f4d"),
    "actual_score": ("Actual Score", "\u5b9e\u9645\u6210\u7ee9"),
    "original_score": ("Original Score", "\u539f\u6210\u7ee9"),
    "routed_score": ("Routed (New)", "\u5206\u6d41\uff08\u65b0\uff09"),
    "tcn": ("TCN",),
}


def is_number(value):
    return isinstance(value, (int, float)) and not isinstance(value, bool) and math.isfinite(value)


def as_number(value):
    if is_number(value):
        return value
    if value in (None, ""):
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    if not math.isfinite(number):
        return None
    return int(number) if number.is_integer() else number


def level_for(actual):
    if actual <= 100:
        return "Low 0-100"
    if actual <= 129:
        return "Intermediate 101-129"
    return "High 130+"


def summarize_accuracy(rows, key):
    valid = [row for row in rows if is_number(row.get(f"{key}AbsError"))]
    if not valid:
        return {
            "n": 0,
            "ok10": 0,
            "acc10": 0,
            "ok5": 0,
            "acc5": 0,
            "mae": 0,
            "medianAbs": 0,
            "bias": 0,
            "maxAbs": 0,
            "over10": 0,
            "under10": 0,
            "severe30": 0,
        }
    errors = [row[f"{key}Error"] for row in valid]
    abs_errors = [abs(error) for error in errors]
    ok10 = sum(value <= 10 for value in abs_errors)
    ok5 = sum(value <= 5 for value in abs_errors)
    return {
        "n": len(valid),
        "ok10": ok10,
        "acc10": ok10 / len(valid),
        "ok5": ok5,
        "acc5": ok5 / len(valid),
        "mae": sum(abs_errors) / len(abs_errors),
        "medianAbs": median(abs_errors),
        "bias": sum(errors) / len(errors),
        "maxAbs": max(abs_errors),
        "over10": sum(error > 10 for error in errors),
        "under10": sum(error < -10 for error in errors),
        "severe30": sum(value > 30 for value in abs_errors),
    }


def summarize_diff(rows):
    valid = [row for row in rows if is_number(row.get("diff"))]
    if not valid:
        return {
            "n": 0,
            "same": 0,
            "newHigher": 0,
            "newLower": 0,
            "avgDiff": 0,
            "medianDiff": 0,
            "avgAbsDiff": 0,
            "maxAbsDiff": 0,
            "gt10": 0,
            "gt20": 0,
            "gt30": 0,
        }
    diffs = [row["diff"] for row in valid]
    abs_diffs = [abs(diff) for diff in diffs]
    return {
        "n": len(valid),
        "same": sum(diff == 0 for diff in diffs),
        "newHigher": sum(diff > 0 for diff in diffs),
        "newLower": sum(diff < 0 for diff in diffs),
        "avgDiff": sum(diffs) / len(diffs),
        "medianDiff": median(diffs),
        "avgAbsDiff": sum(abs_diffs) / len(abs_diffs),
        "maxAbsDiff": max(abs_diffs),
        "gt10": sum(diff > 10 for diff in abs_diffs),
        "gt20": sum(diff > 20 for diff in abs_diffs),
        "gt30": sum(diff > 30 for diff in abs_diffs),
    }


def group_summary(rows, field):
    result = []
    values = sorted({row.get(field) for row in rows}, key=lambda item: str(item))
    for value in values:
        group = [row for row in rows if row.get(field) == value]
        result.append(
            {
                "group": value,
                "n": len(group),
                "original": summarize_accuracy(group, "original"),
                "new": summarize_accuracy(group, "new"),
                "tcnPeak": summarize_accuracy(group, "tcnPeak"),
                "diff": summarize_diff(group),
            }
        )
    return result


def video_sort_key(item):
    area, video, lane, file_name = item[0]
    digits = "".join(ch for ch in str(file_name) if ch.isdigit())
    return (str(area), int(digits or 0), str(video), str(lane))


def video_summary(rows):
    groups = defaultdict(list)
    for row in rows:
        groups[(row["area"], row["video"], row["lane"], row["file"])].append(row)
    result = []
    for (area, video, lane, file_name), group in sorted(groups.items(), key=video_sort_key):
        result.append(
            {
                "area": area,
                "video": video,
                "lane": lane,
                "file": file_name,
                "n": len(group),
                "actualN": sum(is_number(row.get("actual")) for row in group),
                "original": summarize_accuracy(group, "original"),
                "new": summarize_accuracy(group, "new"),
                "tcnPeak": summarize_accuracy(group, "tcnPeak"),
                "diff": summarize_diff(group),
            }
        )
    return result


def histogram(values, bins):
    counts = []
    for label, lower, upper in bins:
        if upper is None:
            count = sum(value >= lower for value in values)
        else:
            count = sum(lower <= value < upper for value in values)
        counts.append({"label": label, "count": count})
    return counts


def build_record(row_index, row, header):
    def value(name):
        return row[header[name]]

    actual = as_number(value("actual_score"))
    original = as_number(value("original_score"))
    routed = as_number(value("routed_score"))
    tcn = as_number(value("tcn"))
    file_name = str(value("file_name") or "")
    video = str(value("video_id") or file_name.replace(".mkv", ""))
    zone = as_number(value("position"))
    record = {
        "id": f"venue-2-{row_index}-{zone}",
        "area": VENUE_2,
        "row": as_number(value("excel_row")) or row_index,
        "file": file_name,
        "video": video,
        "lane": "2",
        "zone": zone,
        "name": "",
        "school": "",
        "original": original,
        "new": routed,
        "tcnPeak": tcn,
        "actual": actual,
    }
    if is_number(original) and is_number(routed):
        record["diff"] = routed - original
        record["absDiff"] = abs(record["diff"])
    if is_number(routed) and is_number(tcn):
        record["tcnPeakDiff"] = tcn - routed
        record["tcnPeakAbsDiff"] = abs(record["tcnPeakDiff"])
    if is_number(actual):
        record["level"] = level_for(actual)
        for key in ("original", "new", "tcnPeak"):
            if is_number(record.get(key)):
                error = record[key] - actual
                record[f"{key}Error"] = error
                record[f"{key}AbsError"] = abs(error)
                record[f"{key}Acc10"] = abs(error) <= 10
                record[f"{key}Acc5"] = abs(error) <= 5
    return record


def load_records():
    workbook = openpyxl.load_workbook(WORKBOOK, data_only=True)
    sheet_name = next((name for name in SHEET_ALIASES if name in workbook.sheetnames), None)
    if sheet_name is None:
        expected = " or ".join(repr(name) for name in SHEET_ALIASES)
        raise ValueError(f"Workbook is missing the expected sheet: {expected}")
    sheet = workbook[sheet_name]
    header_values = [cell.value for cell in sheet[1]]
    source_header = {name: index for index, name in enumerate(header_values) if name}
    header = {
        key: next((source_header[alias] for alias in aliases if alias in source_header), None)
        for key, aliases in COLUMN_ALIASES.items()
    }
    missing = [key for key, index in header.items() if index is None]
    if missing:
        raise ValueError(f"Position comparison sheet is missing columns: {', '.join(sorted(missing))}")

    records = []
    for row_index, row_values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(value is not None for value in row_values):
            continue
        record = build_record(row_index, row_values, header)
        if record["area"] == VENUE_2 and is_number(record.get("zone")):
            records.append(record)
    workbook.close()
    return records


def main():
    records = load_records()
    actual_rows = [row for row in records if is_number(row.get("actual"))]
    output_rows = [row for row in records if any(is_number(row.get(key)) for key in ("original", "new", "tcnPeak"))]
    tcn_rows = [row for row in actual_rows if is_number(row.get("tcnPeak"))]
    status_counts = Counter((row.get("newAcc10"), row.get("tcnPeakAcc10")) for row in tcn_rows)
    output_abs_diffs = [abs(row["diff"]) for row in output_rows if is_number(row.get("diff"))]

    data = {
        "meta": {
            "sourceWorkbook": str(WORKBOOK),
            "generatedAt": "2026-07-02",
            "threshold": THRESHOLD,
            "scope": "Venue 2 only; Venue 1 is training data and is excluded from evaluation",
            "actualRows": len(actual_rows),
            "outputRows": len(output_rows),
            "tcnPeakRows": len(tcn_rows),
        },
        "actualRows": actual_rows,
        "outputRows": output_rows,
        "summaries": {
            "actual": {
                "original": summarize_accuracy(actual_rows, "original"),
                "new": summarize_accuracy(actual_rows, "new"),
                "tcnPeak": summarize_accuracy(actual_rows, "tcnPeak"),
                "tcnPeakStatus": {
                    "bothCorrect": status_counts[(True, True)],
                    "newWrongTcnPeakCorrect": status_counts[(False, True)],
                    "newCorrectTcnPeakWrong": status_counts[(True, False)],
                    "bothWrong": status_counts[(False, False)],
                },
                "tcnPeakPaired": {
                    "original": summarize_accuracy(tcn_rows, "original"),
                    "new": summarize_accuracy(tcn_rows, "new"),
                    "tcnPeak": summarize_accuracy(tcn_rows, "tcnPeak"),
                },
                "byArea": group_summary(actual_rows, "area"),
                "byLevel": group_summary(actual_rows, "level"),
                "byZone": group_summary(actual_rows, "zone"),
                "byVideo": video_summary(actual_rows),
            },
            "output": {
                "diff": summarize_diff(output_rows),
                "byArea": group_summary(output_rows, "area"),
                "byZone": group_summary(output_rows, "zone"),
                "byVideo": video_summary(output_rows),
                "histogram": histogram(
                    output_abs_diffs,
                    [
                        ("0", 0, 1),
                        ("1-5", 1, 6),
                        ("6-10", 6, 11),
                        ("11-20", 11, 21),
                        ("21-30", 21, 31),
                        (">30", 31, None),
                    ],
                ),
            },
        },
    }
    OUT.write_text("window.DASHBOARD_DATA = " + json.dumps(data, ensure_ascii=True, indent=2) + ";\n", encoding="utf-8")
    print(f"Wrote {OUT}")
    print(f"actualRows={len(actual_rows)} outputRows={len(output_rows)}")


if __name__ == "__main__":
    main()
